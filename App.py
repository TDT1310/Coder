#!/usr/bin/env python
"""
App module for financial fraud detection using Beneish M-Score and Benford's Law.
"""

import re
import io
import base64
import json
import tempfile
from pathlib import Path
from io import BytesIO

import pandas as pd
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
import openpyxl
import markdown

from flask import Flask, request, render_template, redirect, url_for, session
from flask_session import Session
from account_mapping_utils import setup_account_mapping
from AI_model import prepare_excel, rag

matplotlib.use("Agg")  # Use non-GUI backend for server

app = Flask(__name__)
app.secret_key = "your_secret_key"
app.config["SESSION_TYPE"] = "filesystem"
Session(app)

MAPPING_PATH = Path(__file__).parent / "Data" / "account_mapping.xlsx"
standard_account_map, reverse_map, all_std_normalized, all_variant_normalized = (
    setup_account_mapping(MAPPING_PATH, sheet_name=0)
)
# Data extract

def dedup_names(names):
    """Ensure unique names by appending suffix to duplicates."""
    counts = {}
    result = []
    for name in names:
        if name not in counts:
            counts[name] = 0
            result.append(name)
        else:
            counts[name] += 1
            result.append(f"{name}.{counts[name]}")
    return result

def data_extract(uploaded_file, sheet_name):
    """
    Extract financial data from Excel, detect headers, bold formatting,
    and generate unique full account labels.
    """
    # Step 1: Detect header row (where the most year-like values are)
    data = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=None)
    data.fillna(0, inplace=True)

    def is_year(s):
        if isinstance(s, (np.integer, int, float, np.floating)):
            s = str(int(s))
        if isinstance(s, str):
            return bool(re.fullmatch(r"20\d{2}|19\d{2}", s.strip()))
        return False

    max_year_count = 0
    header_row_index = 0
    for i in range(min(20, len(data))):
        year_count = sum(is_year(cell) for cell in data.iloc[i])
        if year_count > max_year_count:
            max_year_count = year_count
            header_row_index = i

    # Step 2: Read again with detected header row
    df = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=header_row_index)
    df = df.dropna(axis=1, how="all").dropna(axis=0, how="all")
    df = df.loc[:, ~df.columns.duplicated()].reset_index(drop=True)
    df.columns = [str(col).strip() for col in df.columns]

    # Step 3: Find account column (auto-detect by string ratio)
    def ratio_string(col):
        non_null = col.dropna()
        return non_null.apply(lambda x: isinstance(x, str)).sum() / len(non_null) if len(non_null) else 0

    ratios = {col: ratio_string(df[col]) for col in df.columns}
    account_col = sorted(ratios.items(), key=lambda x: x[1], reverse=True)[0][0]
    account_col_idx = df.columns.get_loc(account_col) + 1

    # Step 4: Get IsBold info for the account column (aligning to actual data table)
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    ws = wb[sheet_name]
    first_data_row = header_row_index + 2
    is_bold = [
        row[0].font.bold if row[0].font else False
        for row in ws.iter_rows(min_row=first_data_row, max_row=ws.max_row,
                                min_col=account_col_idx, max_col=account_col_idx)
    ]
    is_bold = is_bold[: len(df)]
    df["IsBold"] = is_bold

    # Step 5: Build Full Account with parent context and sheet name
    full_accounts = []
    parent_stack = []
    for _, row in df.iterrows():
        label = str(row[account_col]).strip()
        bold = row["IsBold"]
        if bold:
            parent_stack = [label]
            full_account = f"{sheet_name} - {label}"
        else:
            full_account = f"{sheet_name} - {' - '.join(parent_stack)} - {label}" if parent_stack else f"{sheet_name} - {label}"
        full_accounts.append(re.sub(r"\s+", " ", full_account))

    df["Full Account"] = dedup_names(full_accounts)
    return df.set_index("Full Account")

def transformer(uploaded_file):
    """Merge all sheets from the uploaded Excel file into a standardized DataFrame."""
    excel_file = pd.ExcelFile(uploaded_file)
    sheets = excel_file.sheet_names
    combined_data = [data_extract(uploaded_file, sheet) for sheet in sheets]
    final_data = pd.concat(combined_data, ignore_index=False)
    final_data.index = final_data.index.str.lower()
    return final_data

# Function to calculate all M-score inputs and result
def extract_year_columns(df):
    """Return columns that are likely year values."""
    return [
        col for col in df.columns
        if isinstance(col, int)
        or (isinstance(col, str) and col.strip().isdigit() and len(col.strip()) == 4)
    ]

# --- Beneish M-Score ---
def compute_m_score_components(df):
    """Compute Beneish M-Score components for each period."""
    data_dict = df.T.to_dict()
    year_columns = extract_year_columns(df)
    results = []

    def get(item, year):
        return data_dict[item][str(year)]

    for i in range(len(year_columns) - 1):
        y1, y2 = year_columns[i], year_columns[i + 1]

        try:
            dsri = (
                get("b·∫£ng c√¢n ƒë·ªëi k·∫ø to√°n - c√°c kho·∫£n ph·∫£i thu", y2)
                / get("k·∫øt qu·∫£ kinh doanh - doanh thu thu·∫ßn", y2)
            ) / (
                get("b·∫£ng c√¢n ƒë·ªëi k·∫ø to√°n - c√°c kho·∫£n ph·∫£i thu", y1)
                / get("k·∫øt qu·∫£ kinh doanh - doanh thu thu·∫ßn", y1)
            )

            gm_t = (
                get("k·∫øt qu·∫£ kinh doanh - doanh thu thu·∫ßn", y2)
                - get("k·∫øt qu·∫£ kinh doanh - gi√° v·ªën h√†ng b√°n", y2)
            ) / get("k·∫øt qu·∫£ kinh doanh - doanh thu thu·∫ßn", y2)

            gm_t1 = (
                get("k·∫øt qu·∫£ kinh doanh - doanh thu thu·∫ßn", y1)
                - get("k·∫øt qu·∫£ kinh doanh - gi√° v·ªën h√†ng b√°n", y1)
            ) / get("k·∫øt qu·∫£ kinh doanh - doanh thu thu·∫ßn", y1)

            gmi = gm_t1 / gm_t

            aqi = (
                1
                - (
                    get("b·∫£ng c√¢n ƒë·ªëi k·∫ø to√°n - t√†i s·∫£n ng·∫Øn h·∫°n", y2)
                    + get("b·∫£ng c√¢n ƒë·ªëi k·∫ø to√°n - t√†i s·∫£n c·ªë ƒë·ªãnh", y2)
                )
                / get("b·∫£ng c√¢n ƒë·ªëi k·∫ø to√°n - t·ªïng c·ªông t√†i s·∫£n", y2)
            ) / (
                1
                - (
                    get("b·∫£ng c√¢n ƒë·ªëi k·∫ø to√°n - t√†i s·∫£n ng·∫Øn h·∫°n", y1)
                    + get("b·∫£ng c√¢n ƒë·ªëi k·∫ø to√°n - t√†i s·∫£n c·ªë ƒë·ªãnh", y1)
                )
                / get("b·∫£ng c√¢n ƒë·ªëi k·∫ø to√°n - t·ªïng c·ªông t√†i s·∫£n", y1)
            )

            sgi = get("k·∫øt qu·∫£ kinh doanh - doanh thu thu·∫ßn", y2) / get(
                "k·∫øt qu·∫£ kinh doanh - doanh thu thu·∫ßn", y1
            )

            depi = (
                get("l∆∞u chuy·ªÉn ti·ªÅn t·ªá - kh·∫•u hao tscƒë v√† bƒësƒët", y1)
                / (
                    get("l∆∞u chuy·ªÉn ti·ªÅn t·ªá - kh·∫•u hao tscƒë v√† bƒësƒët", y1)
                    + get("b·∫£ng c√¢n ƒë·ªëi k·∫ø to√°n - t√†i s·∫£n c·ªë ƒë·ªãnh", y1)
                )
            ) / (
                get("l∆∞u chuy·ªÉn ti·ªÅn t·ªá - kh·∫•u hao tscƒë v√† bƒësƒët", y2)
                / (
                    get("l∆∞u chuy·ªÉn ti·ªÅn t·ªá - kh·∫•u hao tscƒë v√† bƒësƒët", y2)
                    + get("b·∫£ng c√¢n ƒë·ªëi k·∫ø to√°n - t√†i s·∫£n c·ªë ƒë·ªãnh", y2)
                )
            )

            sgai = (
                get("k·∫øt qu·∫£ kinh doanh - chi ph√≠ qu·∫£n l√Ω doanh nghi·ªáp", y2)
                / get("k·∫øt qu·∫£ kinh doanh - doanh thu thu·∫ßn", y2)
            ) / (
                get("k·∫øt qu·∫£ kinh doanh - chi ph√≠ qu·∫£n l√Ω doanh nghi·ªáp", y1)
                / get("k·∫øt qu·∫£ kinh doanh - doanh thu thu·∫ßn", y1)
            )

            lvgi = (
                get("b·∫£ng c√¢n ƒë·ªëi k·∫ø to√°n - n·ª£ ph·∫£i tr·∫£", y2)
                / get("b·∫£ng c√¢n ƒë·ªëi k·∫ø to√°n - t·ªïng c·ªông t√†i s·∫£n", y2)
            ) / (
                get("b·∫£ng c√¢n ƒë·ªëi k·∫ø to√°n - n·ª£ ph·∫£i tr·∫£", y1)
                / get("b·∫£ng c√¢n ƒë·ªëi k·∫ø to√°n - t·ªïng c·ªông t√†i s·∫£n", y1)
            )

            tata = (
                get("k·∫øt qu·∫£ kinh doanh - l√£i/(l·ªó) thu·∫ßn sau thu·∫ø", y2)
                - get("l∆∞u chuy·ªÉn ti·ªÅn t·ªá - l∆∞u chuy·ªÉn ti·ªÅn t·ªá r√≤ng t·ª´ c√°c ho·∫°t ƒë·ªông s·∫£n xu·∫•t kinh doanh", y2)
            ) / get("b·∫£ng c√¢n ƒë·ªëi k·∫ø to√°n - t·ªïng c·ªông t√†i s·∫£n", y2)

            m_score = (-4.84 + 0.92 * dsri + 0.528 * gmi + 0.404 * aqi + 0.892 * sgi +
                       0.115 * depi - 0.172 * sgai + 4.679 * tata - 0.327 * lvgi)

            results.append({
                "Period": f"{y1}‚ûû{y2}",
                "DSRI": round(dsri, 4),
                "GMI": round(gmi, 4),
                "AQI": round(aqi, 4),
                "SGI": round(sgi, 4),
                "DEPI": round(depi, 4),
                "SGAI": round(sgai, 4),
                "LVGI": round(lvgi, 4),
                "TATA": round(tata, 4),
                "M-Score": round(m_score, 4)
            })
        except (KeyError, ZeroDivisionError, TypeError):
            continue

    return results

# --- Benford's Distribution ---
def compute_benford_all_periods(df):
    # Filter by IsBold == True
    bold_df = df
    year_columns = extract_year_columns(df)
    benford_results = {}

    def leading_digit(value):
        while value < 1 and value != 0:
            value *= 10
        return int(str(abs(value))[0]) if value != 0 else 0

    for i in range(len(year_columns) - 1):
        y1, y2 = str(year_columns[i]), str(year_columns[i + 1])
        period = f"{y1}‚ûû{y2}"
        values = pd.concat([bold_df[y1], bold_df[y2]]).values.flatten()
        values = [v for v in values if isinstance(v, (float, int)) and v > 0]
        digits = [leading_digit(v) for v in values]
        actual_counts = pd.Series(digits).value_counts().sort_index()
        actual_percentages = actual_counts / actual_counts.sum() * 100

        benford_dist = {d: np.log10(1 + 1 / d) * 100 for d in range(1, 10)}
        benford_df = pd.Series(benford_dist)

        comparison_df = pd.DataFrame({
            "Benford (%)": benford_df,
            "Actual (%)": actual_percentages
        }).fillna(0)

        mad = np.mean(np.abs(comparison_df["Actual (%)"] - comparison_df["Benford (%)"]))

        benford_results[period] = {
            "comparison_df": comparison_df,
            "mad": round(mad, 4)
        }

    return bold_df, benford_results

# ƒê√ÇY L√Ä PH·∫¶N GIAO DI·ªÜN
def fig_to_base64(fig):
    """Convert a Matplotlib figure to base64-encoded PNG image string."""
    buffer = BytesIO()
    fig.savefig(buffer, format="png", bbox_inches="tight")
    buffer.seek(0)
    img_base64 = base64.b64encode(buffer.read()).decode("utf-8")
    plt.close(fig)
    return img_base64

def mscore_line_chart(m_score_table):
    """Generate M-Score line chart and return base64 image string."""
    df = pd.DataFrame(m_score_table)
    fig, ax = plt.subplots(figsize=(6, 3))
    df.set_index("Period")["M-Score"].plot(ax=ax, marker="o", color="#1d4ed8", linewidth=2)
    ax.set_title("M-Score Over Time", color="#1d4ed8")
    ax.set_ylabel("M-Score", color="#1d4ed8")
    ax.set_xlabel("Period", color="#1d4ed8")
    ax.grid(True, color="#fef9c3", alpha=0.5)
    ax.tick_params(axis="x", colors="#4f5871")
    ax.tick_params(axis="y", colors="#4f5871")
    fig.patch.set_facecolor("#fff")
    return fig_to_base64(fig)

def benford_bar_chart(comparison_df, period):
    """Generate Benford comparison bar chart and return base64 image string."""
    fig, ax = plt.subplots(figsize=(6, 3))
    comparison_df.plot(
        kind="bar",
        ax=ax,
        color={"Benford (%)": "#ec4899", "Actual (%)": "#be1862"},
        edgecolor="#fff"
    )
    ax.set_xlabel("Leading Digit", color="#1d4ed8")
    ax.set_ylabel("Percentage", color="#1d4ed8")
    ax.set_title(f"Benford's Law vs Actual ({period})", color="#ec4899")
    ax.grid(True, color="#fef9c3", alpha=0.5)
    ax.tick_params(axis="x", colors="#4f5871")
    ax.tick_params(axis="y", colors="#4f5871")
    fig.patch.set_facecolor("#fff")
    return fig_to_base64(fig)

m_score_explanations = {
    "DSRI": (
        "A significant increase in receivables relative to sales may indicate premature revenue "
        "recognition to artificially boost earnings."
    ),
    "GMI": (
        "A declining gross margin may signal deteriorating business performance, prompting firms "
        "to manipulate profits."
    ),
    "AQI": (
        "An increase in long-term assets‚Äîexcluding property, plant and equipment‚Äîrelative to total "
        "assets suggests aggressive capitalization of costs, potentially inflating earnings."
    ),
    "SGI": (
        "While high growth does not imply manipulation, rapidly expanding firms may face greater "
        "pressure to meet market expectations, increasing the temptation to adjust reported earnings."
    ),
    "DEPI": (
        "A decline in depreciation expense relative to net fixed assets may reflect changes in "
        "accounting estimates that increase reported income."
    ),
    "SGAI": (
        "A disproportionate rise in SG&A expenses compared to sales can be viewed negatively by "
        "analysts, incentivizing management to adjust earnings figures."
    ),
    "LVGI": (
        "An increase in leverage (total debt relative to total assets) can pressure firms to "
        "manipulate earnings to comply with debt covenants."
    ),
    "TATA": (
        "Higher accruals indicate greater use of discretionary accounting practices, which may be "
        "associated with earnings manipulation."
    )
}

def top_mscore_changes(m_score_table):
    """
    Returns a dictionary with period as key and list of top 3 changing variables as values.
    """
    variables = ["DSRI", "GMI", "AQI", "SGI", "DEPI", "SGAI", "LVGI", "TATA"]
    top_changes_by_period = {}

    for i in range(1, len(m_score_table)):
        curr = m_score_table[i]
        prev = m_score_table[i - 1]
        period = curr["Period"]

        deltas = []
        for var in variables:
            try:
                delta = curr[var] - prev[var]
                explanation = m_score_explanations.get(var, "")
                deltas.append((var, curr[var], delta, explanation))
            except Exception:
                continue

        top3 = sorted(deltas, key=lambda x: abs(x[2]), reverse=True)[:3]
        top_changes_by_period[period] = top3

    return top_changes_by_period

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        file = request.files.get("file-upload")
        if file:
            try:
                # Read file into memory
                file_bytes = file.read()
                uploaded_file = io.BytesIO(file_bytes)
                # Process the file to get combined data
                final_data = transformer(uploaded_file)

                # Reset pointer for further reading
                uploaded_file.seek(0)
                # Save to a temp file for prepare_excel
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                    tmp.write(file_bytes)
                    tmp_path = tmp.name
                session["excel_file_path"] = tmp_path
                # Store data as JSON for dashboard use
                session["final_data"] = final_data.to_json()
                # Compute M-score and interpretation for dashboard
                m_score_table = compute_m_score_components(final_data)
                session["m_score_table"] = m_score_table  # Store the table for chart
                session["top_mscore_changes"] = top_mscore_changes(m_score_table)
                if m_score_table:
                    session["m_score"] = m_score_table[-1]["M-Score"]
                else:
                    session["m_score"] = None
                # Compute Benford results and store as JSON-serializable
                _, benford_results = compute_benford_all_periods(final_data)
                # Convert DataFrames to dict for JSON serialization
                benford_results_serializable = {}
                for period, res in benford_results.items():
                    benford_results_serializable[period] = {
                        "comparison_df": res["comparison_df"].to_dict(orient="index"),
                        "mad": res["mad"]
                    }
                session["benford_results"] = benford_results_serializable
                # Interpretation (optional)
                session["analysis_result"] = "See dashboard for details"
                # Redirect to dashboard to display results
                return redirect(url_for("dashboard"))
            except Exception as e:
                err_msg = f"L·ªói khi x·ª≠ l√Ω file: {e}"
                print("Error:", err_msg)
                return render_template("upload.html", error=err_msg)
    return render_template("upload.html")

@app.route("/dashboard", methods=["GET", "POST"])
def dashboard():
    if "analysis_result" not in session:
        return redirect(url_for("index"))
    final_data = pd.read_json(session["final_data"])
    m_score_table = session.get("m_score_table", [])
    benford_results = session.get("benford_results", {})

    # Prepare periods for dashboard
    year_columns = extract_year_columns(final_data)
    periods = []
    for i in range(len(year_columns) - 1):
        y1, y2 = year_columns[i], year_columns[i + 1]
        periods.append(f"{y1}-{y2}")

    # Get selected period from GET/POST param, default to most recent
    selected_period = request.values.get("selected_period")
    if not selected_period or selected_period not in periods:
        selected_period = periods[-1] if periods else None

    top_changes_by_period = session.get("top_mscore_changes", {})
    top_changes = top_changes_by_period.get(selected_period.replace("-", "‚ûû"), [])

    # Find the corresponding M-Score value for the selected period
    m_score_value = None
    for row in m_score_table:
        if row.get("Period") == selected_period.replace("-", "‚ûû"):
            m_score_value = row.get("M-Score")
            break

    # Prepare M-Score Plotly data (all periods for the line chart)
    mscore_plotly_data = {}
    if m_score_table:
        mscore_plotly_data = {
            "x": [row["Period"] for row in m_score_table],
            "y": [row["M-Score"] for row in m_score_table]
        }

    # Prepare Benford chart and MAD for the selected period
    mad = None
    plotly_benford_data = {}
    benford_key = selected_period.replace("-", "‚ûû") if selected_period else None
    if benford_key and benford_key in benford_results:
        comparison_df = pd.DataFrame.from_dict(benford_results[benford_key]["comparison_df"], orient="index")
        mad = benford_results[benford_key]["mad"]
        plotly_benford_data = {
            "x": list(comparison_df.index),
            "benford": list(comparison_df["Benford (%)"]),
            "actual": list(comparison_df["Actual (%)"])
        }
    else:
        mad = None
        plotly_benford_data = {}

    # Prepare Beneish M-Score component bar chart data for selected period
    mscore_components_bar_data = {}
    variables = ["DSRI", "GMI", "AQI", "SGI", "DEPI", "SGAI", "LVGI", "TATA"]
    selected_period_label = selected_period.replace("-", "‚ûû") if selected_period else None
    idx = None
    for i, row in enumerate(m_score_table):
        if row.get("Period") == selected_period_label:
            idx = i
            break
    if idx is not None:
        current = m_score_table[idx]
        previous = m_score_table[idx - 1] if idx > 0 else None
        mscore_components_bar_data = {
            "variables": variables,
            "current": [current[var] for var in variables],
            "current_label": selected_period_label.split("‚ûû")[1] if selected_period_label and "‚ûû" in selected_period_label else "T",
            "previous": [previous[var] for var in variables] if previous else [None]*len(variables),
            "previous_label": selected_period_label.split("‚ûû")[0] if selected_period_label and "‚ûû" in selected_period_label else "T-1"
        }
    # --- Q&A Block ---
    qa_history = session.get("qa_history", [])
    latest_answer = session.get("latest_answer", "")
    latest_qa = session.get("latest_qa", {})

    rag_results_by_period = session.get("rag_results_by_period", {})
    rag_conclusion = None
    if selected_period in rag_results_by_period:
        rag_conclusion = rag_results_by_period[selected_period].get("conclusion")

    return render_template(
        "dashboard.html",
        result=session.get("analysis_result"),
        m_score=m_score_value,
        mad=mad,
        plotly_benford_data=json.dumps(plotly_benford_data),
        mscore_plotly_data=json.dumps(mscore_plotly_data),
        periods=periods,
        selected_period=selected_period,
        mscore_components_bar_data=json.dumps(mscore_components_bar_data),
        top_changes=top_changes,
        qa_history=qa_history,
        latest_answer=latest_answer,
        latest_qa=latest_qa,
        rag_conclusion=rag_conclusion,
    )

@app.route("/dashboard_analysis", methods=["GET"])
def dashboard_analysis():
    final_data = pd.read_json(session["final_data"])
    m_score_table = session.get("m_score_table", [])
    benford_results = session.get("benford_results", {})

    # Prepare periods for dashboard
    year_columns = extract_year_columns(final_data)
    periods = []
    for i in range(len(year_columns) - 1):
        y1, y2 = year_columns[i], year_columns[i + 1]
        periods.append(f"{y1}-{y2}")

    selected_period = request.args.get("selected_period")
    if not selected_period or selected_period not in periods:
        selected_period = periods[-1] if periods else None

    top_changes_by_period = session.get("top_mscore_changes", {})
    top_changes = top_changes_by_period.get(selected_period.replace("-", "‚ûû"), [])

    # Find the corresponding M-Score value for the selected period
    m_score_value = None
    for row in m_score_table:
        if row.get("Period") == selected_period.replace("-", "‚ûû"):
            m_score_value = row.get("M-Score")
            break

    # Prepare M-Score Plotly data (all periods for the line chart)
    mscore_plotly_data = {}
    if m_score_table:
        mscore_plotly_data = {
            "x": [row["Period"] for row in m_score_table],
            "y": [row["M-Score"] for row in m_score_table]
        }

    # Prepare Benford chart and MAD for the selected period
    mad = None
    plotly_benford_data = {}
    benford_key = selected_period.replace("-", "‚ûû") if selected_period else None
    if benford_key and benford_key in benford_results:
        comparison_df = pd.DataFrame.from_dict(benford_results[benford_key]["comparison_df"], orient="index")
        mad = benford_results[benford_key]["mad"]
        plotly_benford_data = {
            "x": list(comparison_df.index),
            "benford": list(comparison_df["Benford (%)"]),
            "actual": list(comparison_df["Actual (%)"])
        }
    else:
        mad = None
        plotly_benford_data = {}

    # Prepare Beneish M-Score component bar chart data for selected period
    mscore_components_bar_data = {}
    variables = ["DSRI", "GMI", "AQI", "SGI", "DEPI", "SGAI", "LVGI", "TATA"]
    selected_period_label = selected_period.replace("-", "‚ûû") if selected_period else None
    idx = None
    for i, row in enumerate(m_score_table):
        if row.get("Period") == selected_period_label:
            idx = i
            break
    if idx is not None:
        current = m_score_table[idx]
        previous = m_score_table[idx - 1] if idx > 0 else None
        mscore_components_bar_data = {
            "variables": variables,
            "current": [current[var] for var in variables],
            "current_label": selected_period_label.split("‚ûû")[1] if selected_period_label and "‚ûû" in selected_period_label else "T",
            "previous": [previous[var] for var in variables] if previous else [None]*len(variables),
            "previous_label": selected_period_label.split("‚ûû")[0] if selected_period_label and "‚ûû" in selected_period_label else "T-1"
        }

    return render_template(
        "dashboard_analysis_partial.html",
        m_score=m_score_value,
        mad=mad,
        plotly_benford_data=json.dumps(plotly_benford_data),
        mscore_plotly_data=json.dumps(mscore_plotly_data),
        mscore_components_bar_data=json.dumps(mscore_components_bar_data),
        top_changes=top_changes,
        selected_period=selected_period,
        periods=periods,
    )

@app.route("/ask", methods=["POST"])
def ask():
    question = request.form.get("user_question")
    excel_file_path = session.get("excel_file_path")
    excel_df = prepare_excel(excel_file_path) if excel_file_path else None

    # Indicate that the AI is generating an answer
    session["latest_answer"] = "Generating answer..."

    answer = rag(excel_df, question) if excel_df else "Knowledge base not available."
    answer_html = markdown.markdown(answer, extensions=["fenced_code", "tables", "nl2br"])

    # Save latest Q&A to session
    session["latest_qa"] = {"question": question, "answer": answer_html}
    session["latest_answer"] = answer_html  # Update with the actual answer

    return answer_html  # Return only the answer HTML

@app.route("/rag_risk_analysis", methods=["POST"])
def rag_risk_analysis():
    selected_period = request.form.get("selected_period")
    m_score_table = session.get("m_score_table", [])
    benford_results = session.get("benford_results", {})
    m_score = None
    mad = None
    # Find M-Score for selected period
    if selected_period and m_score_table:
        for row in m_score_table:
            if row.get("Period") == selected_period.replace("-", "‚ûû"):
                m_score = row.get("M-Score")
                break
    # Find MAD for selected period
    if selected_period and benford_results:
        benford_key = selected_period.replace("-", "‚ûû")
        if benford_key in benford_results:
            mad = benford_results[benford_key]["mad"]
    # Extract the two years from selected_period
    years = selected_period.split("-") if selected_period and "-" in selected_period else []
    year1, year2 = years[0], years[1] if len(years) == 2 else ("", "")
    excel_file_path = session.get("excel_file_path")
    excel_db = prepare_excel(excel_file_path) if excel_file_path else None
    output_retrieval_merged = ""
    if excel_db:
        # Use the same retrieval as in rag() for context
        retrieved_docs = excel_db.similarity_search("financial statement manipulation", k=5)
        output_retrieval_merged = "\n".join([doc.page_content for doc in retrieved_docs])
    # Compose the detailed prompt with wrappers for M-Score and MAD and the two years
    prompt = f"""
You are a financial forensic analyst AI. Based on the financial context below, evaluate the likelihood of financial statement manipulation using the following three methods for the period {year1} and {year2}:

--------------------
üìÑ Context:
{output_retrieval_merged}
--------------------

<<M_SCORE>>{m_score}<</M_SCORE>>
<<MAD>>{mad}<</MAD>>

üéØ Your task is to:
1. Compute and assign a score for each of the 3 fraud detection components:
   - **Beneish M-Score**
   - **Benford MAD**
   - **Red Flag Count**
   Each is scored from 0 to 2 based on the following rules:

üîπ **Scoring Criteria**:
- **M-Score**:
  - ‚â§ -2.22 ‚Üí 2 points (no alert)
  - -2.22 < M ‚â§ -1.78 ‚Üí 1 point (mild warning)
  - > -1.78 ‚Üí 0 points (strong warning)

- **Benford MAD**:
  - < 0.006 ‚Üí 2 points
  - 0.006 ‚â§ MAD < 0.012 ‚Üí 1 point
  - ‚â• 0.012 ‚Üí 0 points

- **Red Flags** (each count as 1):
  - < 2 red flags ‚Üí 1 point
  - ‚â• 2 red flags ‚Üí 0 points

üîç Red flags to detect (explain if present, only use data from {year1} and {year2}):
- CFO / Net Income < 0 or < 0.5 for 2 consecutive years
- Other Receivables / Total Receivables > 0.5
- Bad Debt Provision / Receivables > 0.3
- CFO is negative while Net Income is positive for 2 years
- Revenue growth > 50% YoY but CFO doesn't grow accordingly

--------------------
‚úÖ Your output should include:
1. A total score out of 6
2. A brief summary for:
   - The M-Score value, and its risk level
   - The Benford MAD value and its deviation level
   - The number of red flags and what they are
3. A conclusion with a risk interpretation:
   - 1 point: Extremely high risk of manipulation
   - 2 points: Very high risk of manipulation
   - 3 points: High risk of manipulation
   - 4 points: Slight risk - Needs attention
   - 5 points: Low risk - No action needed
   - 6 points: Very low risk - No action needed
4. how each of the red flags are calculated and their values and their formulas

If the context lacks enough data, respond with:  
**"I do not have enough financial data to compute a fraud risk score."**  and what data is lacking
"""
    if excel_db:
        rag_result = rag(excel_db, prompt)
    else:
        rag_result = "Knowledge base not available."
    rag_result_html = markdown.markdown(rag_result, extensions=["fenced_code", "tables", "nl2br"])
    # Extract the conclusion (must be one of the 6 allowed strings)
    allowed_conclusions = [
        "Extremely high risk of manipulation",
        "Very high risk of manipulation",
        "High risk of manipulation",
        "Slight risk - Needs attention",
        "Low risk - No action needed",
        "Very low risk - No action needed"
    ]
    conclusion = None
    # Try to find an allowed conclusion in the HTML (bolded or not)
    for allowed in allowed_conclusions:
        if allowed in rag_result_html:
            conclusion = allowed
            break
    if not conclusion:
        # Fallback: search in plain text result
        for allowed in allowed_conclusions:
            if allowed in rag_result:
                conclusion = allowed
                break
    # Store the conclusion for this period in a session dictionary
    rag_results_by_period = session.get("rag_results_by_period", {})
    rag_results_by_period[selected_period] = {
        "conclusion": conclusion,
        "full_result": rag_result_html
    }
    session["rag_results_by_period"] = rag_results_by_period
    return rag_result_html

@app.route("/recommendations", methods=["POST"])
def recommendations():
    selected_period = request.form.get("selected_period")
    # Extract the two years from selected_period
    years = selected_period.split("-") if selected_period and "-" in selected_period else []
    year1, year2 = years[0], years[1] if len(years) == 2 else ("", "")
    excel_file_path = session.get("excel_file_path")
    excel_db = prepare_excel(excel_file_path) if excel_file_path else None
    output_retrieval_merged = ""
    if excel_db:
        # Use the same retrieval as in rag() for context
        retrieved_docs = excel_db.similarity_search("financial statement manipulation", k=5)
        output_retrieval_merged = "\n".join([doc.page_content for doc in retrieved_docs])
    prompt = f"""
You are a financial forensic analyst AI.  
Based on the financial context below and detected red flags in the company's financial statements from {year1} until {year2} to generate clear, actionable recommendations for areas that should be further investigated before making an investment decision.

--------------------
üìÑ Context:
{output_retrieval_merged}
--------------------

Red Flags to be Detected (only use data from {year1} and {year2}):
- CFO / Net Income < 0 or < 0.5 for 2 consecutive years
- Other Receivables / Total Receivables > 0.5
- Bad Debt Provision / Receivables > 0.3
- CFO is negative while Net Income is positive for 2 years
- Revenue growth > 50% YoY but CFO doesn't grow accordingly

üéØ For each red flag:
- Identify which section(s) of the financial statements or disclosures (e.g., revenue recognition, cash flow statement, accounts receivable, related party transactions, management discussion, etc.) are most relevant to the red flag.
- Show the value for the accounts used
- Explain why this area requires further attention, referencing the nature of the red flag.
- Suggest specific questions the investor should ask or evidence/disclosures they should review to assess the true risk or to clarify the anomaly.
If a red flag is particularly severe or raises suspicion of fraud, clearly recommend consulting auditors or seeking independent verification.

Format your response as a **numbered list** with a brief summary at the end.  
If there are no red flags, simply state:  
"No material red flags were detected; no further investigation is recommended."

"""
    excel_file_path = session.get("excel_file_path")
    excel_db = prepare_excel(excel_file_path) if excel_file_path else None
    if excel_db:
        recommendations_result = rag(excel_db, prompt)
    else:
        recommendations_result = "Knowledge base not available."
    recommendations_result_html = markdown.markdown(recommendations_result, extensions=["fenced_code", "tables", "nl2br"])
    return recommendations_result_html

@app.route("/summary_card", methods=["POST"])
def summary_card():
    selected_period = request.form.get("selected_period")
    rag_results_by_period = session.get("rag_results_by_period", {})
    rag_conclusion = None
    if selected_period in rag_results_by_period:
        rag_conclusion = rag_results_by_period[selected_period].get("conclusion")
    return render_template("summary_card.html", rag_conclusion=rag_conclusion)

@app.route("/reset_analysis", methods=["POST"])
def reset_analysis():
    keys_to_clear = [
        "final_data", "m_score_table", "benford_results", "analysis_result", "m_score", "top_mscore_changes",
        "rag_results_by_period", "qa_history", "latest_answer", "latest_qa", "excel_file_path"
    ]
    for key in keys_to_clear:
        session.pop(key, None)
    return redirect(url_for("index"))

if __name__ == "__main__":
    app.run(debug=True)
