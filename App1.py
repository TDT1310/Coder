# ĐÂY LÀ PHẦN KHAI BÁO THƯ VIỆN
from account_mapping_utils import setup_account_mapping, robust_get
from flask import Flask, request, render_template, redirect, url_for, session
from flask_session import Session
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import re
import io
import openpyxl
from pathlib import Path
import streamlit as st

app = Flask(__name__)
app.secret_key = "your_secret_key" 
app.config["SESSION_TYPE"] = "filesystem"  
Session(app)

# Standardize the path for account mapping
MAPPING_PATH = Path(__file__).parent / "Data" / "account_mapping.xlsx"
standard_account_map, reverse_map, all_std_normalized, all_variant_normalized = setup_account_mapping(MAPPING_PATH, sheet_name=0)

# DEF MỘT SỐ HÀM
# Data extract

def dedup_names(names):
    counts = {}
    result = []
    for name in names:
        if name not in counts:
            counts[name] = 0
            result.append(name)
        else:
            counts[name] += 1
            result.append(f"{name}.{counts[name]}")
    return result

def Data_extract(uploaded_file, sheet_name):

    """
    Extracts the financial data table from an Excel sheet, detects header,
    aligns with bold formatting, and builds a unique Full Account label.
    """
    # Step 1: Detect header row (where the most year-like values are)
    data = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=None)
    data.fillna(0, inplace=True)

    def is_year(s):
        if isinstance(s, (np.int64, np.float64, int, float)):
            s = str(int(s))
        if isinstance(s, str):
            return bool(re.fullmatch(r'20\d{2}|19\d{2}', s.strip()))
        return False

    max_year_count = 0
    header_row_index = 0
    for i in range(min(20, len(data))):  # Scan first 20 rows
        year_count = sum(is_year(cell) for cell in data.iloc[i])
        if year_count > max_year_count:
            max_year_count = year_count
            header_row_index = i

    # Step 2: Read again with detected header row
    df = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=header_row_index)
    df = df.dropna(axis=1, how='all')
    df = df.dropna(axis=0, how='all')
    df = df.loc[:, ~df.columns.duplicated()]
    df = df.reset_index(drop=True)
    df.columns = [str(col).strip() for col in df.columns]

    # Step 3: Find account column (auto-detect by string ratio)
    def ratio_string(col):
        non_null = col.dropna()
        if len(non_null) == 0:
            return 0
        num_str = non_null.apply(lambda x: isinstance(x, str)).sum()
        return num_str / len(non_null)

    ratios = {col: ratio_string(df[col]) for col in df.columns}
    ratios_sorted = sorted(ratios.items(), key=lambda x: x[1], reverse=True)
    account_col = ratios_sorted[0][0]
    account_col_idx = df.columns.get_loc(account_col)+1

    # Step 4: Get IsBold info for the account column (aligning to actual data table)
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    ws = wb[sheet_name]
    # openpyxl is 1-based, pandas is 0-based (+2 = header + 1)
    first_data_row = header_row_index + 2
    is_bold = []
    for idx, row in enumerate(ws.iter_rows(min_row=first_data_row, max_row=ws.max_row, min_col=account_col_idx, max_col=account_col_idx)):
        cell = row[0]
        is_bold.append(cell.font.bold if cell.font else False)
    is_bold = is_bold[:len(df)]
    df['IsBold'] = is_bold

    # Step 5: Build Full Account with parent context and sheet name
    full_accounts = []
    parent_stack = []
    for idx, row in df.iterrows():
        label = str(row[account_col]).strip()
        bold = row['IsBold']
        if bold:
            parent_stack = [label]
            full_account = f"{sheet_name} - {label}"
        else:
            if parent_stack:
                full_account = f"{sheet_name} - {' - '.join(parent_stack)} - {label}"
            else:
                full_account = f"{sheet_name} - {label}"
        full_account = re.sub(r'\s+', ' ', full_account)
        full_accounts.append(full_account)
    df['Full Account'] = full_accounts
    df['Full Account'] = dedup_names(df['Full Account'])
    df = df.set_index('Full Account')
    return df

# Chuẩn hóa data
def transformer (uploaded_file):
     # Tiến hành imput file để phân tích 
    excel_file = pd.ExcelFile(uploaded_file)
    sheets = excel_file.sheet_names
    combined_data = []
    for names in sheets:
        data_test = Data_extract(uploaded_file, names)
        combined_data.append(data_test)
    #Kết hợp và ghi vào file excel mới
    final_data = pd.concat(combined_data, ignore_index=False)
    final_data.index = final_data.index.str.lower()
    final_data.to_excel("final_data.xlsx", index=True)
    return final_data

def year(df):
    return [col for col in df.columns if isinstance(col, int) or (isinstance(col, str) and col.strip().isdigit() and len(col.strip()) == 4)]

# --- Beneish M-Score ---
def compute_m_score_components(df):
    data_dict = df.T.to_dict()
    year_columns = year(df)
    results = []

    def get(item, year): return data_dict[item][str(year)]

    for i in range(len(year_columns) - 1):
        y1, y2 = year_columns[i], year_columns[i + 1]
        try:
            dsri = (get("bảng cân đối kế toán - các khoản phải thu", y2) / get("kết quả kinh doanh - doanh thu thuần", y2)) / (get("bảng cân đối kế toán - các khoản phải thu", y1) / get("kết quả kinh doanh - doanh thu thuần", y1))
            gm_t = (get("kết quả kinh doanh - doanh thu thuần", y2) - get("kết quả kinh doanh - giá vốn hàng bán", y2)) / get("kết quả kinh doanh - doanh thu thuần", y2)
            gm_t1 = (get("kết quả kinh doanh - doanh thu thuần", y1) - get("kết quả kinh doanh - giá vốn hàng bán", y1)) / get("kết quả kinh doanh - doanh thu thuần", y1)
            gmi = gm_t1 / gm_t
            aqi = (1 - (get("bảng cân đối kế toán - tài sản ngắn hạn", y2) + get("bảng cân đối kế toán - tài sản cố định", y2)) / get("bảng cân đối kế toán - tổng cộng tài sản", y2)) / (1 - (get("bảng cân đối kế toán - tài sản ngắn hạn", y1) + get("bảng cân đối kế toán - tài sản cố định", y1)) / get("bảng cân đối kế toán - tổng cộng tài sản", y1))
            sgi = get("kết quả kinh doanh - doanh thu thuần", y2) / get("kết quả kinh doanh - doanh thu thuần", y1)
            depi = (get("thuyết minh - chi phí sản xuất theo yếu tố - chi phí khấu hao tài sản cố định", y1) / (get("thuyết minh - chi phí sản xuất theo yếu tố - chi phí khấu hao tài sản cố định", y1) + get("bảng cân đối kế toán - tài sản cố định", y1))) / (get("thuyết minh - chi phí sản xuất theo yếu tố - chi phí khấu hao tài sản cố định", y2) / (get("thuyết minh - chi phí sản xuất theo yếu tố - chi phí khấu hao tài sản cố định", y2) + get("bảng cân đối kế toán - tài sản cố định", y2)))
            sgai = (get("kết quả kinh doanh - chi phí quản lý doanh nghiệp", y2) / get("kết quả kinh doanh - doanh thu thuần", y2)) / (get("kết quả kinh doanh - chi phí quản lý doanh nghiệp", y1) / get("kết quả kinh doanh - doanh thu thuần", y1))
            lvgi = (get("bảng cân đối kế toán - nợ phải trả", y2) / get("bảng cân đối kế toán - tổng cộng tài sản", y2)) / (get("bảng cân đối kế toán - nợ phải trả", y1) / get("bảng cân đối kế toán - tổng cộng tài sản", y1))
            tata = (get("kết quả kinh doanh - lãi/(lỗ) thuần sau thuế", y2) - get("lưu chuyển tiền tệ - lưu chuyển tiền tệ ròng từ các hoạt động sản xuất kinh doanh", y2)) / get("bảng cân đối kế toán - tổng cộng tài sản", y2)

            m_score = -4.84 + 0.92*dsri + 0.528*gmi + 0.404*aqi + 0.892*sgi + \
                       0.115*depi - 0.172*sgai + 4.679*tata - 0.327*lvgi

            results.append({
                "Period": f"{y1}➞{y2}",
                "DSRI": round(dsri, 4), "GMI": round(gmi, 4), "AQI": round(aqi, 4),
                "SGI": round(sgi, 4), "DEPI": round(depi, 4), "SGAI": round(sgai, 4),
                "LVGI": round(lvgi, 4), "TATA": round(tata, 4), "M-Score": round(m_score, 4)
            })
        except Exception:
            continue

    return results

# --- Benford Analysis ---
def compute_benford_all_periods(df):
    # Filter by IsBold == True
    bold_df = df[df['IsBold'] == True]

    year_columns = year(df)
    benford_results = {}

    def leading_digit(v):
        while v < 1:
            v *= 10
        return int(str(v)[0])

    for i in range(len(year_columns) - 1):
        y1 = str(year_columns[i])
        y2 = str(year_columns[i + 1])
        period = f"{y1}➞{y2}"

        values = pd.concat([bold_df[y1], bold_df[y2]]).values.flatten()
        values = [v for v in values if isinstance(v, (float, int)) and v > 0]
        leading_digits = [leading_digit(v) for v in values]

        actual_counts = pd.Series(leading_digits).value_counts().sort_index()
        actual_percentages = actual_counts / actual_counts.sum() * 100

        benford_dist = {d: np.log10(1 + 1/d) * 100 for d in range(1, 10)}
        benford_df = pd.Series(benford_dist)

        comparison_df = pd.DataFrame({
            'Benford (%)': benford_df,
            'Actual (%)': actual_percentages
        }).fillna(0)

        mad = np.mean(np.abs(comparison_df['Actual (%)'] - comparison_df['Benford (%)']))

        benford_results[period] = {
            "comparison_df": comparison_df,
            "mad": round(mad, 4)
        }

    return bold_df, benford_results


st.title ("Ứng dụng phát hiện gian lận báo cáo tài chính")
st.markdown ("""Vui lòng tải lên báo cáo tài chính dưới định dạng CSV hoặc Excel để phân tích.""")

# Upload tài liệu và phân tích

uploaded_file = st.file_uploader("Tải lên báo cáo tài chính",type=["xlsx", "csv"])
if uploaded_file is not None:
    final_data = transformer(uploaded_file)
    st.write(final_data)
    bold_df, benford_results = compute_benford_all_periods(final_data)
    st.write(bold_df)

    st.subheader("M-Score Chi tiết theo từng giai đoạn")
    m_score_table = compute_m_score_components(final_data)
    m_score_df = pd.DataFrame(m_score_table)
    st.line_chart(m_score_df.set_index("Period")[["M-Score"]])

    # Detect year columns
    year_options = year(final_data)
    year_pairs = [(year_options[i], year_options[i+1]) for i in range(len(year_options)-1)]

    # Let user choose a year pair
    selected_pair = st.selectbox("Chọn giai đoạn phân tích:", year_pairs, format_func=lambda x: f"{x[0]} ➞ {x[1]}")
    y1, y2 = selected_pair

    # Collapsed sections for both analyses
    with st.expander("Kết quả Beneish M-Score"):
        selected_period = f"{y1}➞{y2}"
        row_index = m_score_df.index[m_score_df["Period"] == selected_period].tolist()

        idx = row_index[0]
        current = m_score_df.loc[idx]
        previous = m_score_df.loc[idx - 1] if idx > 0 else None

        # Compute and sort deltas
        variables = ["DSRI", "GMI", "AQI", "SGI", "DEPI", "SGAI", "LVGI", "TATA"]
        deltas = []

        if idx > 0:
            previous = m_score_df.loc[idx - 1]

            # Bar chart: compare actual values for 8 variables in T-1 vs T
            st.markdown("### 📊 So sánh giá trị các biến Beneish giữa hai kỳ")

            var_values = {
                'Variable': variables,
                'T-1': [previous[var] for var in variables],
                'T': [current[var] for var in variables]
            }
            var_df = pd.DataFrame(var_values)

            fig, ax = plt.subplots(figsize=(10, 4))
            width = 0.35
            x = np.arange(len(variables))

            ax.bar(x - width/2, var_df['T-1'], width, label=f"{y1}")
            ax.bar(x + width/2, var_df['T'], width, label=f"{y2}")
            ax.set_xticks(x)
            ax.set_xticklabels(var_df['Variable'], rotation=45)
            ax.set_ylabel("Giá trị biến")
            ax.set_title(f"So sánh biến Beneish: {selected_period}")
            ax.legend()
            st.pyplot(fig)



        for var in variables:
            if previous is not None:
                delta = current[var] - previous[var]
                deltas.append((var, current[var], delta))
            else:
                deltas.append((var, current[var], None))

        # Sort by absolute delta and keep top 3
        top_changes = sorted(deltas, key=lambda x: abs(x[2]) if x[2] is not None else 0, reverse=True)[:3]

        # Build formatted output
        def format_var(name, value, delta):
            if delta is None:
                return f"{name}: {value:.4f}"
            color = "green" if delta > 0 else "red" if delta < 0 else "gray"
            sign = "+" if delta > 0 else ""
            return f'{name}: {value:.4f} <span style="color:{color}; font-size: 0.9em">({sign}{delta:.4f})</span>'

        # Format Markdown
        html_output = f"""
        <h4>M-score Analysis ({selected_period})</h4>
        <ul>
        {''.join(f"<li>{format_var(var, val, delta)}</li>" for var, val, delta in top_changes)}
        </ul>
        <h4><b>M-Score</b>: <code>{current['M-Score']:.4f}</code></h4>
        """

        st.markdown(html_output, unsafe_allow_html=True)

        score = current['M-Score']
        if score < -2.22:
            st.success("Unlikely Manipulation")
        elif score < -1.78:
            st.warning("Possible Manipulation")
        else:
            st.error("Likely Manipulation")

    with st.expander("Phân tích Benford's Law"):
        period_key = f"{y1}➞{y2}"
        bdata = benford_results[period_key]
        comparison_df = bdata["comparison_df"]
        mad = bdata["mad"]

        st.subheader(f"Benford Analysis ({period_key})")
        st.markdown("Comparison Table")
        st.dataframe(comparison_df.style.format("{:.2f}"))
        st.markdown("Bar Chart")
        fig, ax = plt.subplots(figsize=(10, 5))
        comparison_df.plot(kind='bar', ax=ax)
        ax.set_xlabel("Leading Digit")
        ax.set_ylabel("Percentage")
        ax.set_title(f"Benford's Law vs Actual ({period_key})")
        ax.grid(True)
        st.pyplot(fig)
        st.markdown("MAD (Mean Absolute Deviation) Test")
        st.markdown(f"**MAD:** `{mad:.4f}`")

        # Interpretation
        if mad <= 0.006:
            st.success("✅ Close conformity with Benford's Law")
        elif mad <= 0.012:
            st.info("🟡 Acceptable conformity")
        else:
            st.error("❌ Nonconformity — possible anomaly")

# ĐÂY LÀ PHẦN GIAO DIỆN

#@app.route("/", methods=["GET", "POST"])
#def index():
#    if request.method == "POST":
#        file = request.files.get("file-upload")
#        print("file uploased:", file)
#        if file:
#            try:
#                print("file is not empty")
#                # Đọc file về bộ nhớ tạm
#                file_bytes = file.read()
#                uploaded_file = io.BytesIO(file_bytes)
#                # Chạy transformer để lấy data tổng hợp
#                final_data = transformer(uploaded_file)
#                # Reset lại io cho Data_extract vì pandas cần đọc lại file từ đầu
#                uploaded_file.seek(0)
#                extracted_data = Data_extract(uploaded_file, "Thuyết minh")
#                # Store only the data as JSON or CSV, not HTML
#                print("final_data index:", list(final_data.index))
#                session["final_data"] = final_data.to_json()
#                session["extracted_data"] = extracted_data.to_json()
#                session["analysis_result"] = compute_m_score_components(final_data)
#                print("redirecting to dashboard")
#                return redirect(url_for("dashboard"))
#            
#            except Exception as e:
#                err_msg = f"Lỗi khi xử lý file: {e}"
#                print("Error:", err_msg)
#                return render_template("upload.html", error=err_msg)
#
#    return render_template("upload.html")

#@app.route("/dashboard")
#def dashboard():
#    if "analysis_result" not in session:
#        return redirect(url_for("index"))
#    # Convert JSON back to DataFrame
#    table_html = pd.read_json(session["final_data"]).to_html(classes="table table-striped", border=0)
#    extracted_html = pd.read_json(session["extracted_data"]).to_html(classes="table table-striped", border=0)
#    return render_template(
#       "dashboard.html",
#        result=session.get("analysis_result"),
#        table_html=table_html,
#        extracted_html=extracted_html
#    )

#if __name__ == "__main__":
#    app.run(debug=True)